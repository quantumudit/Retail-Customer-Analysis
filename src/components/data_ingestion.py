"""
summary
"""

import csv
import re
from os.path import dirname, exists, join, normpath

import httpx
from openpyxl import load_workbook

from src.constants import CONFIGS
from src.exception import CustomException
from src.logger import logger
from src.utils.basic_utils import create_directories, read_yaml, unzip_file


class DataIngestion:
    """
    summary
    """

    def __init__(self):
        """
        Initializes the DataIngestion class. Reads the configuration files.
        """
        # Read the configuration files
        self.configs = read_yaml(CONFIGS).data_ingestion

        # Source data configs
        self.data_url = self.configs.data_url
        self.user_agent = self.configs.user_agent
        self.timeout = self.configs.timeout
        self.download_status = self.configs.download_status

        # Output directories & filepath
        self.external_filepath = normpath(self.configs.external_path)
        self.raw_dir = normpath(self.configs.raw_dir)
        self.interim_dir = normpath(self.configs.interim_dir)

    def download_data(self):
        """_summary_

        Raises:
            CustomException: _description_
        """
        try:
            if not exists(self.external_filepath) or self.download_status:
                # Create directory if not exist
                create_directories([dirname(self.external_filepath)])

                # Construct header
                headers = {"User-Agent": self.user_agent,
                           "accept-language": "en-US"}

                # Download and save the data
                logger.info("File download started")
                with httpx.stream(
                    "GET", self.data_url, headers=headers, timeout=self.timeout
                ) as response:
                    with open(self.external_filepath, "wb") as file:
                        for chunk in response.iter_bytes():
                            file.write(chunk)
                logger.info("File downloaded successfully")
            else:
                logger.info(
                    "The %s already exists. Skipping download", self.external_filepath
                )
        except Exception as e:
            logger.error(CustomException(e))
            raise CustomException(e) from e

    def save_dataset(self):
        """_summary_

        Raises:
            CustomException: _description_
        """
        try:
            # Create directory if not exist
            create_directories([self.raw_dir, self.interim_dir])

            # Unzip the file
            logger.info("Unzipping the downloaded file")
            unzipped_files = unzip_file(
                zipfile_path=self.external_filepath, unzip_dir=self.raw_dir
            )
            excel_file = [
                file for file in unzipped_files if file.endswith(".xlsx")][0]
            logger.info("File unzipped. Excel file saved at: %s", excel_file)

            # Load excel file & extract sheet names
            logger.info("Loading the excel file")
            excel_filepath = join(self.raw_dir, excel_file)
            workbook = load_workbook(excel_filepath)
            sheet_names = workbook.sheetnames
            logger.info("Excel file loaded successfully.")
            # Iterate over sheets and save data as CSV files
            for sheet_name in sheet_names:
                logger.info("Writing %s sheet into CSV file", sheet_name)
                csv_filename = re.sub(r"[\s-]", "_", sheet_name.lower())
                csv_filepath = join(self.interim_dir, f"{csv_filename}.csv")

                worksheet = workbook[sheet_name]

                with open(csv_filepath, "w", newline="", encoding="utf-8") as cf:
                    writer = csv.writer(cf)
                    for row in worksheet.iter_rows(values_only=True):
                        writer.writerow(row)
                logger.info("%s Sheet data written successfully", sheet_name)
                logger.info("CSV file saved at: %s", csv_filepath)
        except Exception as e:
            logger.error(CustomException(e))
            raise CustomException(e) from e
